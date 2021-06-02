import os
import sys
import numpy as np
import openpyxl
import seem_lib as sl
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import time

############################################################################################
# General code structure

# 1) Get the inputs' file
# 2) Add psspy to the path
# 3) Setting basic parameters and variables
# 4) Loading data from the excel
# 5) Starting psse
# 6) Settings for plotting
# 7) Again loading data from excel
# 8) Getting gen to check info
# 9) Closing all gens and starting only needed ones
# 10) Closing gen to check
# 10) Run fsnl
# 11) Get needed data about the network (buses numbers, free id's, load buses list, loads S)
# 12) Starting calculations:
# 12.1) If P4PF or PF4P, for every hour:
# 12.1.1) Updating loads and closing checked gen
# 12.1.2) Run fnsl
# 12.1.3) Getting pretest voltages
# 12.1.4) Running P4PF or PF for P
# 12.2) If P vs PF
# 12.2.1) Updating loads to wanted and closing checked gen
# 12.2.2) Run fnsl
# 12.2.3) Getting pretest voltages
# 12.2.4) Running P vs PF
# 13) Deleting the tmp case
# 14) Printing running time
# 15) Printing diagrams
# 16) For every bus:
# 16.1) If not in gen's buses list - these buses weren't checked
# 16.2) Save data
############################################################################################


# 1) Get the inputs' file
try:
    input_excel = None
    cur_dir_dirlist = os.listdir(".")
    # There's input.xlsx is in the folder
    if "input.xlsx" in cur_dir_dirlist:
        input_excel = os.path.join(os.getcwd(), "input.xlsx")
    # There's input in argv
    else:
        input_args = sys.argv
        if len(input_args) > 0:
            # Data in argv might be of xlsx type
            if len(input_args[-1] > 4):
                input_args = input_args[-1]
                if input_excel[-4:] == "xlsx":
                    # Found the input in argv
                    input_excel = os.path.join(os.getcwd(), input_args)
    if input_excel is None:
        raise Exception("Couldn't find an input xlsx file")
        exit(1)
finally:
    print("Inputs' file = " + input_excel)

# 2) Add psspy to the path
sys.path.append(r"C:\Program Files (x86)\PTI\PSSEXplore34\PSSBIN")
os.environ['PATH'] = (r"C:\Program Files (x86)\PTI\PSSEXplore34\PSSBIN;" + os.environ['PATH'])
sys.path.append(r"C:\Program Files (x86)\PTI\PSSEXplore34\PSSPY27")
os.environ['PATH'] = (r"C:\Program Files (x86)\PTI\PSSEXplore34\PSSPY27;" + os.environ['PATH'])
sys.path.append(r"C:\Program Files (x86)\PTI\PSSEXplore34\PSSPY34")
os.environ['PATH'] = (r"C:\Program Files (x86)\PTI\PSSEXplore34\PSSPY34;" + os.environ['PATH'])

# imported libraries for psse
import psspy
from psspy import _i
from psspy import _f
import redirect

redirect.psse2py()

# for running time measurements
start_time = time.clock()

# 3) setting basic parameters and variables
# Default variables
voltage_step = 0.1
max_PF = 1
min_PF = 0.9
PF_loads = 0.92
bins_width_for_set_PF = 0.2
bins_width_for_set_P = 0.01
TOIS = 1  # Take only in service

# 4) Loading data from the excel
# Read the input file
wb = openpyxl.load_workbook(input_excel, data_only=True)
ws = wb.active

# PSS/E Saved case
redirect.psse2py()
sav_dir = str(ws['B15'].value)
sav_file = str(ws['B14'].value)
case = os.path.join(sav_dir, sav_file)

# General settings
# Use an existing generator or an imaginary one.
use_new_gen = not (bool(ws['S4'].value))  # the not is to handle the excel
use_list_of_mach_on = bool(ws['I1'].value)
check_at_some_buses = bool(ws['E1'].value)
verbose = bool(ws['C22'].value)

# Type of checking
check_for_set_PF = bool(ws['C6'].value)
check_for_set_P = bool(ws['C4'].value)
check_for_P_vs_PF = bool(ws['C8'].value)
check_SCC_SCR = bool(ws['C12'].value)

# 5) Starting psse
psspy.psseinit(100)
psspy.case(case)
sl.psse_verbose(False)

# 6) settings for plotting
output_dir = str(ws['B16'].value)

if not os.path.isdir(output_dir):
    os.mkdir(output_dir)
    if verbose:
        print("Output folder was created")
plt.ion()
print_hist = False
print_graph = False
Time_between_plots = 10

# ID = Concatenate gmtime
ID = ""
for i in range(len(time.gmtime())):
    ID = ID + str(time.gmtime()[i])

# 7) Again loading data from excel
# Get the P for the tests or abort the test if the input is not valid
# Value for set P
if check_for_set_P and not check_for_P_vs_PF and not check_for_set_PF:
    max_p_to_check = float(ws['B4'].value)

# Value for P VS PF
elif check_for_P_vs_PF and not check_for_set_P and not check_for_set_PF:
    max_p_to_check = float(ws['B8'].value)
    min_p_to_check = float(ws['D8'].value)
    if min_p_to_check >= max_p_to_check:
        raise Exception("The maximal power to check must be bigger then the smallest")
        exit(1)
# Value for set PF
elif check_for_set_PF and not check_for_set_P and not check_for_P_vs_PF:
    max_p_to_check = float(ws['D6'].value)
else:
    raise Exception("You have to choose only one type of check")
    exit(1)

# Get the PF of the loads
PF_loads = float(ws['X4'].value)

# Get the list of machines in use
if use_list_of_mach_on:
    num_of_machines_in_use = int(ws['L4'].value)
    machines_in_use = np.zeros((num_of_machines_in_use, 2))
    for i in range(num_of_machines_in_use):
        index = 4 + i
        bus_loc = "J" + str(index)
        gen_loc = "K" + str(index)
        machines_in_use[i, 0] = int(ws[bus_loc].value)
        machines_in_use[i, 1] = int(ws[gen_loc].value)
else:
    machines_in_use = None

# Get the list of buses to check at
if check_at_some_buses:
    num_of_busses = int(ws['G4'].value)
    buses_to_check_at = []
    for i in range(num_of_busses):
        index = 4 + i
        location = "F" + str(index)
        buses_to_check_at.append(int(ws[location].value))

else:
    buses_to_check_at = None

if not use_new_gen:
    gen_to_check_bus = int(ws['Q4'].value)
    gen_to_check_id = str(ws['R4'].value)

if check_for_set_PF:
    set_PF = float(ws['B6'].value)
else:
    set_PF = None

# Starting a log file in case of verbose
if verbose:
    sys.stdout = open(output_dir + "\\log file ID = " + str(ID) + ".txt", 'w')

############################################################################################
# Hourly vector
############################################################################################
if check_for_set_PF or check_for_set_P:
    number_of_hours_yearly = int(ws['V4'].value)
    if number_of_hours_yearly == 0:
        raise Exception("There must be at least one hour load")
        exit(1)
    percentage_vec = []
    if number_of_hours_yearly > 0:
        for i in range(number_of_hours_yearly):
            cur_i = 5 + i
            cell = "T" + str(cur_i)
            percentage_vec.append(int(float(ws[cell].value) * 100))

############################################################################################
# Transform all the settings to variables
############################################################################################

# getting info of the generators working in this case
# which are not the swing. can be given by user or taking
# all the generators in the case

# Turn only chosen machines on
if use_list_of_mach_on and len(machines_in_use) > 0:
    machines_in_use = np.char.mod('%d', machines_in_use)
    ones_vec = np.ones((machines_in_use.shape[0], 1), dtype=str)
    machines_in_use = np.append(machines_in_use, ones_vec, axis=1)

# Creating list of machines in use #
elif not use_list_of_mach_on:
    machines_in_use = sl.gen_finder()
else:
    machines_in_use = machines_in_use

# Case of both machines in use and use existing gen
if use_list_of_mach_on and not use_new_gen:
    bus_i = np.where(machines_in_use == str(gen_to_check_bus))
    # Found the bus
    if len(bus_i[0]) > 0:
        bus_i = bus_i[0][0]
        gen_i = np.where(machines_in_use[bus_i] == str(gen_to_check_id))
        # Found the generator
        if len(gen_i[0]) > 0:
            gen_i = gen_i[0][0]
        # Couldn't find the gen
        else:
            gen_i = None
    # Couldn't find the bus
    else:
        gen_i = None
    # Couldn't find the gen to check
    if gen_i is None:
        raise Exception("The generator for checking must be on (part of generators in use list)")
        exit(1)

# 8) Getting gen to check info
# if using an already in system gen,
# getting the wanted gen index at the table
if use_new_gen:
    gen4testing = None
else:
    gen4testing = -1
    index_num = -1
    for gen in machines_in_use:
        index_num = index_num + 1
        if int(gen[0]) == int(gen_to_check_bus) and int(gen[1]) == int(gen_to_check_id):
            gen4testing = index_num

# 9) Closing all gens and starting only needed ones
# Make sure that only needed machines are on by closing
# all and starting only needed ones
all_machine_table = sl.gen_finder()
print("\nShutting down all generators and starting only needed ones")
for gen in all_machine_table:
    if verbose:
        print("\nShutting down gen = " + str(gen[1]) + " at bus = " + str(gen[0]))
    sl.machine_on(int(gen[0]), int(gen[1]), False)

if use_list_of_mach_on:
    for gen in machines_in_use:
        if verbose:
            print("\nStarting gen = " + str(gen[1]) + " at bus = " + str(gen[0]))
        # pe.machine_on(int(gen[0]), int(gen[1]), True)
        sl.machine_on(int(gen[0]), int(gen[1]), True)
else:
    for gen in all_machine_table:
        if verbose:
            print("\nStarting gen = " + str(gen[1]) + " at bus = " + str(gen[0]))
        sl.machine_on(int(gen[0]), int(gen[1]), True)

# 10) Run fsnl
sl.seem_fnsl()

# 11) Get needed data about the network (buses numbers, load buses list, loads S)
# Get buses to check at numbers
if check_at_some_buses:
    bus_numbers = list(buses_to_check_at)
else:
    error_code, [bus_numbers] = psspy.abusint(-1, TOIS, 'NUMBER')
    if error_code > 0:
        print("\nproblem getting buses info, error code is = " + str(error_code))

# info for updating loads by percentage
error_code, [load_buses_list] = psspy.aloadint(-1, 4, 'NUMBER')  # 4 - all loads
if error_code > 0:
    print("\ncan't get loaded buses list, error code is = " + str(error_code))
error_code, [base_MVA_of_loads] = psspy.aloadreal(-1, 4, 'TOTALACT')  # 4 - all loads
if error_code > 0:
    print("\ncan't get loads MVA, error code is = " + str(error_code))

##
# 12) Starting calculations,
# 12.1) If P4PF or PF4P, for every hour:
# 12.1.1) Updating loads and closing checked gen
# 12.1.2) Run fnsl
# 12.1.3) Getting pretest voltages
# 12.1.4) Running P4PF or PF for P

############################################################################################
# starting calculations
############################################################################################
if check_for_set_P or check_for_set_PF:

    # output for diagrams
    num_of_hours = len(percentage_vec)
    data_by_hour_dictionary = {}

    for hour_index in range(num_of_hours):
        print("\nCalculating hour No. " + str(hour_index + 1) + " of " + str(num_of_hours))
        # Updating loads by percentage and recalculating the pretest voltages
        sl.machine_on(int(machines_in_use[gen4testing, 0]), int(machines_in_use[gen4testing, 1]), False)
        sl.load_change_by_percentage_from_base(load_buses_list, base_MVA_of_loads, PF_loads, percentage_vec[hour_index])
        sl.seem_fnsl()
        error_code, [pretest_voltages] = psspy.abusreal(-1, TOIS, 'PU')
        if verbose:
            print("\npretest_voltages after load change is: \n" + str(pretest_voltages))
        if error_code > 0:
            print("\nproblem getting pre_test voltages in PU info, error code is = " + str(error_code))

        # output position's format: (bus number, check number, 0=P of check / 1= PF res)
        # output format = | P | PF |
        if check_for_set_PF and not check_for_set_P:
            output = sl.max_P4PF_1hour(max_p_to_check, machines_in_use, pretest_voltages, set_PF, verbose,
                                       gen4testing, buses_to_check_at, check_SCC_SCR)
        elif check_for_set_P and not check_for_set_PF:
            output = sl.min_PF4P_1hour(max_p_to_check, machines_in_use, pretest_voltages, verbose, gen4testing,
                                       buses_to_check_at, check_SCC_SCR)
        elif check_for_set_P and check_for_set_PF:
            raise Exception("You should choose either set the P or the PF, not both")
            exit(1)
            break

        # adding the results for this hour to the general results list (using dictionary)
        data_by_hour_dictionary[hour_index] = output

# 12.2) If P vs PF
# 12.2.1) Updating loads to wanted and closing checked gen
# 12.2.2) Run fnsl
# 12.2.3) Getting pretest voltages
# 12.2.4) Running P vs PF

elif check_for_P_vs_PF:

    percentage = int(100 * float(ws['B10'].value))
    # Updating loads by percentage and recalculating the pretest voltages
    sl.machine_on(int(machines_in_use[gen4testing, 0]), int(machines_in_use[gen4testing, 1]), False)
    sl.load_change_by_percentage_from_base(load_buses_list, base_MVA_of_loads, PF_loads, percentage)
    sl.seem_fnsl()

    error_code, [pretest_voltages] = psspy.abusreal(-1, TOIS, 'PU')
    if verbose:
        print("\npretest_voltages after load change is: \n" + str(pretest_voltages))
    if error_code > 0:
        print("\nproblem getting pre_test voltages in PU info, error code is = " + str(error_code))

    # output for diagrams
    output = sl.each_P_min_PF(max_p_to_check, min_p_to_check, machines_in_use, pretest_voltages, verbose, gen4testing,
                              buses_to_check_at, check_SCC_SCR)

# 13) Deleting the tmp case
try:
    os.remove("./tmp.sav")
except Exception:
    print("\nCouldn't find or access ./tmp.sav")

# 14) Printing running time
# printing statistics about running time
end_time = time.clock()
print("\nrunning time for " + str(len(bus_numbers)) + " buses is: " + str(end_time - start_time) + " Seconds")
print("\nCreating graphs")

# 15) Printing diagrams
# 16) For every bus:
# 16.1) If not in gen's buses list - these buses weren't checked
# 16.2) Save data
############################################################################################
# Printing diagrams
############################################################################################
buses_skipped = sl.gen_finder(True)  # not checking in gen's buses, so no need to save the info

# in set_P case, for each bus, get number of hours at each min PF and printing
if check_for_set_P:
    bus_mins = {}
    # format[bus index num , min_PF_vectors]
    for bus in range(len(bus_numbers)):
        if not (bool(buses_skipped.size) and str(bus_numbers[bus]) in buses_skipped):
            min_PF_vec = np.zeros(num_of_hours)
            # finding number of hours at each P at a bus.
            for hour_index in range(num_of_hours):
                hour_data = data_by_hour_dictionary[hour_index]
                min_PF_vec[hour_index] = hour_data[bus, 1]
            bus_mins[bus] = min_PF_vec

            # printing the histogram

            fig, ax = plt.subplots()
            bins_amount = int((max_PF - min_PF) / bins_width_for_set_P + 2)
            bins_range = (min_PF - 0.5 * bins_width_for_set_P, max_PF + 0.5 * bins_width_for_set_P)
            counts, bins, patches = ax.hist(bus_mins[bus], bins=bins_amount, range=bins_range, align='mid')
            ax.xaxis.set_major_locator(ticker.MultipleLocator(bins_width_for_set_P))
            ax.yaxis.set_major_locator(ticker.FixedLocator(list(dict.fromkeys(counts))))
            for rect in ax.patches:
                height = int(rect.get_height())
                if height == 0:
                    continue
                ax.annotate(height, xy=(rect.get_x() + rect.get_width() / 2, height),
                            xytext=(0, 5), textcoords='offset points', ha='center', va='bottom')
            label_str = "Histogram for bus " + str(bus_numbers[bus]) + " with P of " + str(max_p_to_check) + \
                        "[MW]. \n bin width is " + str(bins_width_for_set_P) + "\nNumber of hours is " + \
                        str(num_of_hours) + ", found solution for " + \
                        str(int(sum(counts))) + " hours"
            plt.title(label_str)
            plt.xlabel('PF')
            plt.ylabel('Number of hours')
            if print_hist:
                plt.show()
                plt.pause(Time_between_plots)
            else:
                plt.savefig(
                    output_dir + "\\Bus=" + str(bus_numbers[bus]) + "_P=" + str(max_p_to_check) + "_Histogram_ID="
                    + str(ID) + ".png")
            plt.close()

# in set_PF case, for each bus, get number of hours at each max P and printing
if check_for_set_PF:
    bus_maxes = {}
    # format[bus index num , max_P_vectors]
    for bus in range(len(bus_numbers)):
        if not (bool(buses_skipped.size) and str(bus_numbers[bus]) in buses_skipped):
            max_P_vec = np.zeros(num_of_hours)
            # finding number of hours at each P at a bus.
            for hour_index in range(num_of_hours):
                hour_data = data_by_hour_dictionary[hour_index]
                max_p_for_hour = max(hour_data[bus, :])
                max_P_vec[hour_index] = max_p_for_hour
            bus_maxes[bus] = max_P_vec

            # printing the histogram
            fig, ax = plt.subplots()
            bins_amount = int((max_p_to_check - 0) / bins_width_for_set_PF + 1)
            bins_range = (0 - 0.5 * bins_width_for_set_PF, max_p_to_check + 0.5 * bins_width_for_set_PF)
            counts, bins, patches = ax.hist(bus_maxes[bus], bins=bins_amount, range=bins_range, align='mid')
            ax.xaxis.set_major_locator(ticker.MultipleLocator(bins_width_for_set_PF * (1 / bins_width_for_set_PF)))
            ax.xaxis.set_minor_locator(ticker.MultipleLocator(bins_width_for_set_PF))
            ax.yaxis.set_major_locator(ticker.FixedLocator(list(dict.fromkeys(counts))))
            for rect in ax.patches:
                height = int(rect.get_height())
                if height == 0:
                    continue
                ax.annotate(height, xy=(rect.get_x() + rect.get_width() / 2, height),
                            xytext=(0, 5), textcoords='offset points', ha='center', va='bottom')
            label_str = "Histogram for bus " + str(bus_numbers[bus]) + " with PF of " + str(set_PF) + \
                        ". \n bin width is " + str(bins_width_for_set_PF) + "\nNumber of hours is " + \
                        str(num_of_hours) + ", found solution for " + \
                        str(int(sum(counts))) + " hours"
            plt.title(label_str)
            plt.xlabel('P range[MW]')
            plt.ylabel('Number of hours')
            left_edge = max(bins_width_for_set_PF * np.min(np.nonzero(counts)) - 0.6, 0)
            plt.xlim([left_edge - 0.5 * bins_width_for_set_PF,
                     max_p_to_check + 0.5 * bins_width_for_set_PF])
            if print_hist:
                plt.show()
                plt.pause(Time_between_plots)
            else:
                plt.savefig(output_dir + "\\Bus=" + str(bus_numbers[bus]) + "_PF=" + str(set_PF) + "_Histogram_"
                            + "ID=" + str(ID) + ".png")
            plt.close()

# if checking for max PF, plotting the P and PF for each bus for one chosen hour
not_skipped_busses = np.setdiff1d(bus_numbers, buses_skipped)
if not check_for_set_PF and not check_for_set_P:
    for it in range(len(bus_numbers)):
        if bool(buses_skipped.size) and (str(bus_numbers[it]) in buses_skipped):
            continue
        else:
            colors = plt.rcParams["axes.prop_cycle"]()
            fig, axs = plt.subplots()
            fig.suptitle("Min PF for given P at bus = " + str(bus_numbers[it]))
            label_str = "Bus " + str(bus_numbers[it])
            plt.scatter(output[it, :, 0], output[it, :, 1], label=label_str)
            plt.yticks(list(dict.fromkeys(output[it, :, 1])))
            plt.xlim([min_p_to_check - bins_width_for_set_PF, max_p_to_check + bins_width_for_set_PF])
            plt.ylim([min_PF - bins_width_for_set_P, max_PF + bins_width_for_set_P])
            plt.legend()
            plt.grid()
            # naming the axes
            plt.xlabel('P [MW]')
            plt.ylabel('PF')

            if print_graph:
                plt.show()
                plt.pause(Time_between_plots)
                plt.close()
            else:
                plt.savefig(output_dir + "\\Max_P_VS_PF_bus=" + str(bus_numbers[it]) + "_graph_id=" + ID + ".png")
                plt.close()
print("Finished graphs with ID = " + str(ID))

# closing the output file
if verbose:
    sys.stdout.close()
